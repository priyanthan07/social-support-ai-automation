"""Generate sample applicant documents for the end-to-end demo.

For each persona this renders:
    emirates_id.png            (image  -> OCR)
    bank_statement.pdf         (PDF, text + table)
    credit_report.pdf          (PDF, text)
    resume.pdf                 (PDF, text)
    assets_liabilities.xlsx    (Excel, tabular)
    persona.json               (ground-truth for reference / demo autofill)

Some personas contain DELIBERATE inconsistencies (address / income) to
demonstrate the Validation Agent.

Run:  uv run python -m ssa_ml.generate_documents
"""

from __future__ import annotations

import json

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

from ssa_ml import paths

# --------------------------------------------------------------------------- #
# Persona definitions (ground truth). Values map to the model feature schema.  #
# --------------------------------------------------------------------------- #
PERSONAS: list[dict] = [
    {
        "key": "aisha_eligible",
        "full_name": "Aisha Abdullah Al Mansoori",
        "emirates_id": "784-1988-1234567-1",
        "dob": "1988-04-12",
        "nationality": "United Arab Emirates",
        "address_id": "Villa 12, Al Wathba, Abu Dhabi",
        "address_credit": "Villa 12, Al Wathba South, Abu Dhabi",  # minor mismatch
        "employment_status": "unemployed",
        "bank_monthly_credit": 1800,     # small support/irregular income
        "stated_monthly_income": 1800,
        "family_size": 6,
        "num_dependents": 4,
        "family_members": [
            {"name": "Khalid Al Mansoori", "relation": "spouse"},
            {"name": "Mariam Al Mansoori", "relation": "child"},
            {"name": "Sara Al Mansoori", "relation": "child"},
            {"name": "Yousef Al Mansoori", "relation": "child"},
            {"name": "Hamdan Al Mansoori", "relation": "child"},
        ],
        "credit_score": 545,
        "employment_history": [
            ("Al Noor Trading LLC", "Administrative Assistant", "2015 - 2021"),
        ],
        "education": "High School",
        "skills": ["MS Office", "Customer Service", "Arabic/English"],
        "assets": [("Savings Account", "Bank", 22000), ("Gold", "Jewelry", 15000)],
        "liabilities": [("Personal Loan", "Bank", 60000)],
    },
    {
        "key": "omar_noteligible",
        "full_name": "Omar Khalid Al Suwaidi",
        "emirates_id": "784-1985-7654321-2",
        "dob": "1985-09-30",
        "nationality": "United Arab Emirates",
        "address_id": "Apartment 1004, Marina Heights, Dubai",
        "address_credit": "Apartment 1004, Marina Heights, Dubai",
        "employment_status": "employed",
        "bank_monthly_credit": 32000,
        "stated_monthly_income": 28000,   # mismatch vs bank statement
        "family_size": 3,
        "num_dependents": 1,
        "family_members": [
            {"name": "Layla Al Suwaidi", "relation": "spouse"},
            {"name": "Khalid Al Suwaidi", "relation": "child"},
        ],
        "credit_score": 760,
        "employment_history": [
            ("Emirates Global Bank", "Senior Manager", "2016 - Present"),
            ("Gulf Finance", "Analyst", "2010 - 2016"),
        ],
        "education": "Bachelor of Finance",
        "skills": ["Financial Analysis", "Leadership", "Risk Management"],
        "assets": [
            ("Primary Residence", "Property", 1800000),
            ("Investment Portfolio", "Securities", 450000),
            ("Vehicle", "Car", 120000),
        ],
        "liabilities": [("Mortgage", "Bank", 900000)],
    },
    {
        "key": "fatima_borderline",
        "full_name": "Fatima Hassan Al Balushi",
        "emirates_id": "784-1992-2468101-3",
        "dob": "1992-01-25",
        "nationality": "United Arab Emirates",
        "address_id": "Building 7, Al Qusais, Dubai",
        "address_credit": "Building 7, Al Qusais, Dubai",
        "employment_status": "self_employed",
        "bank_monthly_credit": 7200,
        "stated_monthly_income": 7200,
        "family_size": 4,
        "num_dependents": 2,
        "family_members": [
            {"name": "Ahmed Al Balushi", "relation": "spouse"},
            {"name": "Noor Al Balushi", "relation": "child"},
            {"name": "Zayed Al Balushi", "relation": "child"},
        ],
        "credit_score": 640,
        "employment_history": [
            ("Self-employed - Home Bakery", "Owner", "2019 - Present"),
            ("Spinneys", "Sales Associate", "2014 - 2019"),
        ],
        "education": "Diploma in Business",
        "skills": ["Baking", "Small Business", "Social Media Marketing"],
        "assets": [("Savings Account", "Bank", 65000), ("Vehicle", "Car", 45000)],
        "liabilities": [("Car Loan", "Bank", 30000), ("Credit Card", "Bank", 12000)],
    },
]


def _font(size: int) -> ImageFont.FreeTypeFont:
    try:
        return ImageFont.load_default(size=size)
    except Exception:  # pragma: no cover - very old Pillow
        return ImageFont.load_default()


def _make_emirates_id(persona: dict, out_dir) -> None:
    img = Image.new("RGB", (1000, 640), "#eef3fb")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, 1000, 90], fill="#1e3a8a")
    d.text((30, 28), "United Arab Emirates - Identity Card", font=_font(30), fill="white")
    d.rectangle([40, 130, 300, 440], outline="#1e3a8a", width=3)
    d.text((90, 270), "PHOTO", font=_font(28), fill="#94a3b8")

    lines = [
        f"Name: {persona['full_name']}",
        f"ID Number: {persona['emirates_id']}",
        f"Date of Birth: {persona['dob']}",
        f"Nationality: {persona['nationality']}",
        f"Address: {persona['address_id']}",
        "Card Expiry: 2029-04-11",
    ]
    y = 150
    for line in lines:
        d.text((340, y), line, font=_font(28), fill="#0f172a")
        y += 55
    img.save(out_dir / "emirates_id.png")


class _PDF(FPDF):
    def header_title(self, title: str) -> None:
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 12, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(2)

    def line_item(self, text: str, bold: bool = False) -> None:
        self.set_font("Helvetica", "B" if bold else "", 11)
        self.multi_cell(0, 7, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def _make_bank_statement(persona: dict, out_dir) -> None:
    pdf = _PDF()
    pdf.add_page()
    pdf.header_title("Emirates National Bank - Account Statement")
    pdf.line_item(f"Account Holder: {persona['full_name']}")
    pdf.line_item(f"Address: {persona['address_id']}")
    pdf.line_item("Statement Period: last 3 months")
    pdf.ln(4)
    credit = persona["bank_monthly_credit"]
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(60, 8, "Date", border=1)
    pdf.cell(70, 8, "Description", border=1)
    pdf.cell(50, 8, "Amount (AED)", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 11)
    rows = [
        ("2026-04-01", "Salary / Income Credit", f"+{credit:,.0f}"),
        ("2026-04-05", "Rent / Housing", f"-{credit * 0.35:,.0f}"),
        ("2026-04-15", "Groceries & Utilities", f"-{credit * 0.25:,.0f}"),
        ("2026-05-01", "Salary / Income Credit", f"+{credit:,.0f}"),
        ("2026-05-10", "Loan Repayment", f"-{credit * 0.15:,.0f}"),
        ("2026-06-01", "Salary / Income Credit", f"+{credit:,.0f}"),
    ]
    for dt, desc, amt in rows:
        pdf.cell(60, 8, dt, border=1)
        pdf.cell(70, 8, desc, border=1)
        pdf.cell(50, 8, amt, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)
    pdf.line_item(f"Average Monthly Credit: AED {credit:,.0f}", bold=True)
    pdf.output(str(out_dir / "bank_statement.pdf"))


def _make_credit_report(persona: dict, out_dir) -> None:
    pdf = _PDF()
    pdf.add_page()
    pdf.header_title("Al Etihad Credit Bureau - Credit Report")
    pdf.line_item(f"Full Name: {persona['full_name']}")
    pdf.line_item(f"Registered Address: {persona['address_credit']}")
    pdf.line_item(f"Credit Score: {persona['credit_score']}", bold=True)
    pdf.ln(3)
    pdf.line_item("Outstanding Liabilities:", bold=True)
    total = 0
    for name, kind, value in persona["liabilities"]:
        total += value
        pdf.line_item(f"  - {name} ({kind}): AED {value:,.0f}")
    pdf.line_item(f"Total Outstanding: AED {total:,.0f}", bold=True)
    pdf.output(str(out_dir / "credit_report.pdf"))


def _make_resume(persona: dict, out_dir) -> None:
    pdf = _PDF()
    pdf.add_page()
    pdf.header_title(persona["full_name"] + " - Curriculum Vitae")
    pdf.line_item(f"Education: {persona['education']}")
    pdf.ln(2)
    pdf.line_item("Employment History:", bold=True)
    for company, role, period in persona["employment_history"]:
        pdf.line_item(f"  - {role}, {company} ({period})")
    pdf.ln(2)
    pdf.line_item("Skills: " + ", ".join(persona["skills"]))
    pdf.output(str(out_dir / "resume.pdf"))


def _make_assets_excel(persona: dict, out_dir) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AssetsLiabilities"
    ws.append(["Category", "Item", "Type", "Value_AED"])
    for name, kind, value in persona["assets"]:
        ws.append(["Asset", name, kind, value])
    for name, kind, value in persona["liabilities"]:
        ws.append(["Liability", name, kind, value])
    wb.save(out_dir / "assets_liabilities.xlsx")


def generate() -> None:
    paths.ensure_dirs()
    for persona in PERSONAS:
        out_dir = paths.DOCUMENTS_DIR / persona["key"]
        out_dir.mkdir(parents=True, exist_ok=True)
        _make_emirates_id(persona, out_dir)
        _make_bank_statement(persona, out_dir)
        _make_credit_report(persona, out_dir)
        _make_resume(persona, out_dir)
        _make_assets_excel(persona, out_dir)
        (out_dir / "persona.json").write_text(json.dumps(persona, indent=2), encoding="utf-8")
        print(f"Generated documents for {persona['key']} -> {out_dir}")


if __name__ == "__main__":
    generate()
